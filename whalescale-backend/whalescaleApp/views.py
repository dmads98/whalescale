import PIL
import boto3
import json
#import coremltools
from django.contrib import messages
from django.contrib.auth import logout, authenticate
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File


from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from rest_framework.authtoken.models import Token

from whalescaleApp.models import Image, Account, Length, Width, Area, Angle
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ImageSerializer, RegistrationSerializer, AccountSerializer

s3_bucket = 'whalescale'
s3_link = 'https://'+s3_bucket+'.s3.amazonaws.com/'
aws_access_key = 'AKIA5IWS3NKFIPBVANM6'
aws_secret_access_key = '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'

# Author: Thomas Chemmanoor
# Contians all functions necessary for the backend REST API to function



# Testing function to see if this homepage on the backend renders an excel image hosted on s3
@csrf_exempt
def index(request):
    #return HttpResponse("Hello, world. You're at the whaleScale homepage.")
    #print("reached homepage")
    return render(request, 'export.html', {'link': s3_link+'excel_0aWXn2a.jpg'})

# Returns an excel file based on an image ID: Get request
@csrf_exempt
@api_view(['GET'])
def getExcelID(request, pk):
    image = get_object_or_404(Image,id = pk)
    print(image.name)
    response = HttpResponse(content_type='application/ms-excel')

    # decide file name
    response['Content-Disposition'] = 'attachment; filename="Results.xlsx"'

    # creating workbook
    wb = Workbook()

    # adding sheet
    ws = wb.active
    ws.title = "WhaleScale Results"


    # column header names, you can use your own headers here
    ws.cell(row=1, column=1, value="Image Name").font = Font(bold=True)
    ws.cell(row=1, column=2, value=image.name)

    ws.cell(row=2, column=1, value="Focal Length").font = Font(bold=True)
    ws.cell(row=2, column=2, value=image.focalLength)

    ws.cell(row=3, column=1, value="Altitude").font = Font(bold=True)
    ws.cell(row=3, column=2, value=image.altitude)

    ws.cell(row=4, column=1, value="Pixel Dimension").font = Font(bold=True)
    ws.cell(row=4, column=2, value=image.pixelDimension)

    currentRow = 8

    try:
        lengths = Length.objects.filter(image=image)
        #print(lengths.count())
    except ObjectDoesNotExist:
        lengths = None
        #print("IS NONE")

    ws.cell(row=6, column=1, value="Length Measurements").font = Font(bold=True)
    ws.cell(row=7, column=1, value="Name").font = Font(bold=True)
    ws.cell(row=7, column=2, value="Length (m)").font = Font(bold=True)
    ws.cell(row=7, column=3, value="Widths (%)").font = Font(bold=True)

    if lengths is None or lengths.count() == 0:
        ws.cell(row=currentRow, column=1, value="None")
    elif lengths is not None:
        for length in lengths:
            ws.cell(row=currentRow, column=1, value=length.name)
            ws.cell(row=currentRow, column=2, value=length.length)
            widths = Width.objects.filter(length=length)
            currentCol = 3
            for width in widths:
                ws.cell(row=currentRow, column=currentCol, value=width.width)
                currentCol+=1
            currentRow+=1

    try:
        angles = Angle.objects.filter(image=image)
    except ObjectDoesNotExist:
        angles = None

    currentRow+=2
    ws.cell(row=currentRow, column=1, value="Angle Measurements").font = Font(bold=True)
    currentRow+=1
    ws.cell(row=currentRow, column=1, value="Name").font = Font(bold=True)
    ws.cell(row=currentRow, column=2, value="Angle (°)").font = Font(bold=True)
    currentRow += 1

    if angles is None or angles.count() == 0:
        ws.cell(row=currentRow, column=1, value="None")
    elif angles is not None:
        for angle in angles:
            ws.cell(row=currentRow, column=1, value=angle.name)
            ws.cell(row=currentRow, column=2, value=angle.angle)
            currentRow+=1

    try:
        areas = Area.objects.filter(image=image)
    except ObjectDoesNotExist:
        areas = None

    currentRow += 2
    ws.cell(row=currentRow, column=1, value="Area Measurements").font = Font(bold=True)
    currentRow += 1
    ws.cell(row=currentRow, column=1, value="Name").font = Font(bold=True)
    ws.cell(row=currentRow, column=2, value="Area (m²)").font = Font(bold=True)
    currentRow += 1

    if areas is None or areas.count() == 0:
        ws.cell(row=currentRow, column=1, value="None")
    elif areas is not None:
        for area in areas:
            ws.cell(row=currentRow, column=1, value=area.name)
            ws.cell(row=currentRow, column=2, value=area.area)
            currentRow+=1

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(response)
    return response


# Takes 2 images from post request and an image id
# First image is the original image without measurements
# Second image is the image with measurements
# Maps both image files links from s3 to the specific image (based on the id) in the database in the link and measured link data fields
@csrf_exempt
@api_view(['POST'])
def uploadImages(request, pk):
    image = get_object_or_404(Image,id = pk)
    session = boto3.session.Session(aws_access_key_id='AKIA5IWS3NKFIPBVANM6',
                                    aws_secret_access_key='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')
    s3 = session.resource('s3')

    image.link = s3_link + image.name + str(image.id)
    image.measured_link = s3_link + image.name + str(image.id) + "measured"
    
    image.save()
    itemsList = list(request.FILES.items())
    s3.Bucket('whalescale').put_object(Key=image.name + str(image.id), Body=itemsList[0][1])
    s3.Bucket('whalescale').put_object(Key=image.name + str(image.id) + "measured", Body=itemsList[1][1])
    return Response("Uploaded Images to S3 Successfully")


@api_view(['GET'])
def imageList(request):
    if request.user.is_anonymous:
        images = Image.objects.all()
    else:
        images = Image.objects.filter(owner = request.user)
    serializer = ImageSerializer(images, many=True)
    return Response(serializer.data)



# returns all images in the database from the get request
# filters the images based on the user that makes the request
# For example: User A when making this request will only gain access to User A's photos
@api_view(['GET'])
def imageList(request):
    if request.user.is_anonymous:
        images = Image.objects.all()
    else:
        images = Image.objects.filter(owner = request.user)
    serializer = ImageSerializer(images, many=True)
    return Response(serializer.data)

# returns one image from the database through a get request
@api_view(['GET'])
def imageDetail(request, pk):
    image = get_object_or_404(Image,id = pk)
    serializer = ImageSerializer(image, many=False)
    return Response(serializer.data)

# returns one updated image from the database through a put request that takes in an image id and an image instance
@api_view(['PUT'])
def imageUpdate(request,pk):
    image = get_object_or_404(Image,id = pk)
    serializer = ImageSerializer(instance=image,data=request.data)

    if(serializer.is_valid()):
        serializer.save()
    return Response(serializer.data)

# deletes an image from an image id from the database and the s3 bucket
@api_view(['DELETE'])
def imageDelete(request,pk):
    image = get_object_or_404(Image,id = pk)
    session = boto3.session.Session(aws_access_key_id=aws_access_key,
                                    aws_secret_access_key=aws_secret_access_key)
    s3 = session.resource('s3')
    s3.Bucket('whalescale').delete_object(Key=image.file.name)
    image.delete()

    return Response("Image deleted")

# creates an account given an account instance
@api_view(['POST'])
def accountCreate(request):
    serializer = RegistrationSerializer(data=request.data)
    data = {}
    if serializer.is_valid():
        account = serializer.save()
        data['response'] = "Successfully registered a new user."
        data['username'] = account.username
        data['email'] = account.email
    else:
        data = serializer.errors
    return Response(data)


# returns all accounts in the database from the get request
@api_view(['GET'])
def accountList(request):
    accounts = Account.objects.all()
    serializer = AccountSerializer(accounts, many=True)
    return Response(serializer.data)

# returns one account from the database through a get request and an account username
@api_view(['GET'])
def accountDetail(request, username):
    account = get_object_or_404(Account,username = username)
    serializer = AccountSerializer(account, many=False)
    return Response(serializer.data)

# returns one updated account from the database through a put request that takes in an account username and an account instance
@api_view(['PUT'])
def accountUpdate(request, username):
    account = get_object_or_404(Account,username = username)
    serializer = AccountSerializer(instance=account,data=request.data)

    if(serializer.is_valid()):
        serializer.save()
    return Response(serializer.data)

# deletes an account from an account username from the database
@api_view(['DELETE'])
def accountDelete(request,username):
    account = get_object_or_404(Account,username = username)
    account.delete()
    return Response("Account deleted")

# logs account out through get request
@api_view(['GET'])
def log_out(request):
    logout(request)
    return Response("Logout Successful")

#Example request
# {
# "username" : "testaccount",
# "password" : "test",
# "firstName" : "Test",
# "lastName" : "Account"
# }
# creates user account and authentication token for future logins and logs the user into the website
@api_view(['POST'])
def register(request):
    serializer = RegistrationSerializer(data=request.data)
    data = {}
    if serializer.is_valid():
        account = serializer.save()
        data['response'] = "Successfully registered a new user."
        data['username'] = account.username
        token = Token.objects.get(user=account).key
        data['token'] = token
    else:
        data = serializer.errors
    return Response(data)
